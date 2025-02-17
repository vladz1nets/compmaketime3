import pandas as pd
import numpy as np
import time
import os
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import colorsys
import logging

logger = logging.getLogger(__name__)

class ScheduleRec:
    def __init__(self, PartName='', opNum=0, machine='', machineInstance=0,
                 T_start=0.0, T_finish=0.0, fileNumber=0):
        self.PartName = PartName
        self.opNum = opNum
        self.machine = machine
        self.machineInstance = machineInstance
        self.T_start = T_start
        self.T_finish = T_finish
        self.fileNumber = fileNumber

class JobRec:
    def __init__(self, partID=0, batchNo=0, batchSize=0,
                 NextOp=1, readyTime=0.0):
        self.partID = partID
        self.batchNo = batchNo
        self.batchSize = batchSize
        self.NextOp = NextOp
        self.readyTime = readyTime

def ReadBatchData(q_part_file_path):
    logger.info(f"Зчитування даних з файлу Q_part: {q_part_file_path}")
    wsQ = pd.read_excel(q_part_file_path, sheet_name='Q_part', header=0)
    PartNames = wsQ.iloc[:, 0].astype(str).str.strip().str.lower().tolist()
    BatchQty = wsQ.iloc[:, 1].fillna(1).astype(int).tolist()
    BatchVolume = wsQ.iloc[:, 2].fillna(1).astype(int).tolist()
    # Створюємо словники для кількості та обсягу партії
    qty_dict = dict(zip(PartNames, BatchQty))
    volume_dict = dict(zip(PartNames, BatchVolume))
    return qty_dict, volume_dict

def ReadProcessingData(details_file_path_list, PartNames, file_numbers):
    logger.info("Зчитування даних з файлів деталей")
    ProcessingTimeArr = {}
    MachineAssign = {}
    partOpsCount = {}
    partOpFileMapping = {}

    for details_file_path in details_file_path_list:
        logger.info(f"Обробка файлу деталей: {details_file_path}")
        try:
            xls = pd.ExcelFile(details_file_path)
        except Exception as e:
            logger.error(f"Не вдалося відкрити файл {details_file_path}: {e}")
            continue
        file_number = file_numbers.get(details_file_path, 0)
        for sheet_name in xls.sheet_names:
            part_name = sheet_name.strip().lower()
            part_key = f"{part_name}_file{file_number}"  # Унікальний ключ для деталі з файлу
            if part_name not in PartNames:
                logger.warning(f"Деталь '{part_name}' не знайдено в Q_part. Пропускаємо.")
                continue
            wsPart = pd.read_excel(details_file_path, sheet_name=sheet_name, header=None)
            op_index = 0  # Початковий індекс операції для кожної деталі
            for op_row in range(2, wsPart.shape[0]):
                machine_name = str(wsPart.iloc[op_row, 0]).strip()
                if machine_name == '':
                    continue
                op_times = wsPart.iloc[op_row, 1:].tolist()
                for op_time in op_times:
                    if pd.isna(op_time) or op_time == '':
                        op_index += 1
                        continue
                    op_time_value = float(op_time)

                    # Ініціалізуємо списки, якщо необхідно
                    if part_key not in ProcessingTimeArr:
                        ProcessingTimeArr[part_key] = []
                        MachineAssign[part_key] = []
                        partOpsCount[part_key] = 0

                    # Додаємо дані про операцію
                    ProcessingTimeArr[part_key].append(op_time_value)
                    MachineAssign[part_key].append(machine_name)
                    op_total_index = partOpsCount[part_key]

                    # Зберігаємо номер файлу для кожної операції
                    key = (part_key, op_total_index)
                    partOpFileMapping[key] = file_number

                    partOpsCount[part_key] += 1
                    op_index += 1  # Збільшуємо індекс операції
    return ProcessingTimeArr, MachineAssign, partOpsCount, partOpFileMapping

def ReadMachineData(stanok_file_path):
    logger.info(f"Зчитування даних з файлу Stanok: {stanok_file_path}")
    wsStanok = pd.read_excel(stanok_file_path, sheet_name='Stanok', header=0)
    machineAvailability = {}
    machineInstances = []
    for idx, row in wsStanok.iterrows():
        machine_name = str(row[0]).strip()
        if machine_name != '':
            mCount = int(row[1]) if not pd.isna(row[1]) else 1
            mCount = max(mCount, 1)
            machineAvailability[machine_name] = [0.0] * mCount
            for instIdx in range(1, mCount + 1):
                machineInstances.append(f"{machine_name}{instIdx}")
    return machineAvailability, machineInstances

def ComputeDynamicSchedule(qty_dict, volume_dict, ProcessingTimeArr,
                           MachineAssign, partOpsCount, machineAvailability, partOpFileMapping):
    logger.info("Початок розрахунку динамічного розкладу")

    # Отримуємо список унікальних part_keys
    part_keys = list(ProcessingTimeArr.keys())
    partsCount = len(part_keys)

    # Створюємо відображення між part_key і індексом
    partKeyToIndex = {part_key: idx for idx, part_key in enumerate(part_keys)}
    indexToPartKey = {idx: part_key for part_key, idx in partKeyToIndex.items()}

    # Створюємо список черг для кожної деталі
    job_queues = [[] for _ in range(partsCount)]
    part_indices = [0] * partsCount
    total_jobs_remaining = 0

    for idx, part_key in enumerate(part_keys):
        original_part_name = part_key.split('_file')[0]
        qty = qty_dict.get(original_part_name, 0)
        if qty == 0:
            logger.warning(f"Деталь '{original_part_name}' не має кількості у файлі Q_part. Пропускаємо.")
            continue
        volume = volume_dict.get(original_part_name, 1)
        num_batches = qty // volume
        if qty % volume != 0:
            num_batches += 1
        r = qty % volume if qty % volume != 0 else volume
        for batch in range(1, num_batches + 1):
            if (batch == num_batches) and (qty % volume != 0):
                batch_size = r
            else:
                batch_size = volume
            job = JobRec(
                partID=idx,
                batchNo=batch,
                batchSize=batch_size,
                NextOp=1,
                readyTime=0.0
            )
            job_queues[idx].append(job)
            total_jobs_remaining += 1

    scheduleRecArray = []
    machineUsage = {machine: times[:] for machine, times in machineAvailability.items()}

    while total_jobs_remaining > 0:
        for idx in range(partsCount):
            if total_jobs_remaining == 0:
                break

            queue = job_queues[idx]
            if part_indices[idx] >= len(queue):
                continue

            job = queue[part_indices[idx]]
            part_key = indexToPartKey[job.partID]
            opNum = job.NextOp - 1

            if opNum >= len(MachineAssign.get(part_key, [])):
                part_indices[idx] += 1
                total_jobs_remaining -= 1
                continue

            reqMachine = MachineAssign[part_key][opNum]
            procTime_i = job.batchSize * ProcessingTimeArr[part_key][opNum]

            if reqMachine == '':
                candidateStart = job.readyTime
                candidateFinish = candidateStart + procTime_i
                candidateMachineIndex = None
            else:
                availArr = machineUsage[reqMachine]
                earliestTime = min(availArr)
                candidateMachineIndex = availArr.index(earliestTime)
                candidateStart = max(job.readyTime, earliestTime)
                candidateFinish = candidateStart + procTime_i

            finishTime = candidateFinish

            key = (part_key, opNum)
            fileNumber = partOpFileMapping.get(key, 0)

            # Отримуємо оригінальну назву деталі
            original_part_name = part_key.split('_file')[0]

            scheduleRecArray.append(ScheduleRec(
                PartName=f"{original_part_name} (Партія {job.batchNo})",
                opNum=job.NextOp,
                machine=reqMachine,
                machineInstance=candidateMachineIndex + 1 if candidateMachineIndex is not None else 0,
                T_start=candidateStart,
                T_finish=finishTime,
                fileNumber=fileNumber
            ))

            if reqMachine != '':
                machineUsage[reqMachine][candidateMachineIndex] = finishTime

            job.readyTime = finishTime
            job.NextOp += 1

            if job.NextOp > partOpsCount.get(part_key, 0):
                part_indices[idx] += 1
                total_jobs_remaining -= 1
    makespan = max((rec.T_finish for rec in scheduleRecArray), default=0.0)

    return makespan, scheduleRecArray

def DrawGanttChartTable(scheduleArray, makespan, machineInstances, output_filename='GanttChart.xlsx'):
    logger.info("Побудова діаграми Ганта")
    wb = Workbook()
    ws = wb.active
    ws.title = "GanttChart"

    # Обчислюємо максимальну тривалість (округлено до хвилини)
    maxMinutes = int(np.ceil(makespan))

    # Розбиваємо часову ось на інтервали по 10 хвилин
    totalIntervals = maxMinutes // 10
    if maxMinutes % 10 > 0:
        totalIntervals += 1

    # Створення заголовків для годин (кожні 6 інтервалів = 1 година)
    totalHours = int(np.ceil(totalIntervals / 6))
    for hr in range(totalHours):
        colStart = (hr * 6) + 2
        colEnd = colStart + 5
        ws.merge_cells(start_row=1, start_column=colStart, end_row=1, end_column=colEnd)
        cell = ws.cell(row=1, column=colStart)
        cell.value = f"{hr:02d}:00"
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Заголовки для інтервалів по 10 хвилин (другий рядок)
    for minVal in range(0, totalIntervals * 10, 10):
        col = (minVal // 10) + 2
        hourPart = minVal // 60
        minutePart = minVal % 60
        timeStr = f"{hourPart}:{minutePart:02d}"
        cell = ws.cell(row=2, column=col)
        cell.value = timeStr
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Запис імен машинних екземплярів у перший стовпець (починаємо з рядка 3)
    for idx, machine in enumerate(machineInstances):
        ws.cell(row=idx + 3, column=1, value=machine).alignment = Alignment(horizontal='center')

    # Малювання блоків операцій згідно з розкладом
    machineUsage = defaultdict(list)
    color_map = {}

    for rec in scheduleArray:
        if rec.machine == '':
            continue
        machineKey = f"{rec.machine}{rec.machineInstance}"
        if machineKey in machineInstances:
            machineRow = machineInstances.index(machineKey) + 3
        else:
            continue

        startMinute = rec.T_start
        finishMinute = int(np.ceil(rec.T_finish))
        startCol = int(startMinute // 10) + 2
        endCol = int((finishMinute - 1) // 10) + 2
        if startCol > endCol:
            endCol = startCol

        # Генеруємо унікальний колір для кожної деталі з урахуванням номера файлу
        part_identifier = f"{rec.PartName}_{rec.fileNumber}"
        if part_identifier not in color_map:
            hue = hash(part_identifier) % 360
            color = colorsys.hsv_to_rgb(hue / 360, 0.5, 0.8)
            rgb = tuple(int(255 * c) for c in color)
            hex_color = '{:02X}{:02X}{:02X}'.format(*rgb)
            color_map[part_identifier] = hex_color
        else:
            hex_color = color_map[part_identifier]
        fill_color = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

        # Встановлюємо значення комірки з інформацією про операцію
        cell_value = f"{rec.PartName} [Файл {rec.fileNumber}], Оп{rec.opNum}"

        ws.merge_cells(start_row=machineRow, start_column=startCol, end_row=machineRow, end_column=endCol)
        cell = ws.cell(row=machineRow, column=startCol)
        cell.value = cell_value
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = fill_color

        # Додаємо дані для розрахунку простоїв
        machineUsage[machineKey].append((startMinute, finishMinute))

    # Розрахунок та запис простоїв для кожного машинного екземпляра
    for machineKey in machineUsage:
        totalWorkTime = sum(task[1] - task[0] for task in machineUsage[machineKey])
        totalIdleTime = maxMinutes - totalWorkTime
        idlePercentage = (totalIdleTime / maxMinutes) * 100 if maxMinutes > 0 else 0

        machineRow = machineInstances.index(machineKey) + 3
        idleCell = ws.cell(row=machineRow, column=(totalIntervals + 2))
        idleCell.value = f"Простій: {idlePercentage:.2f}%"
        idleCell.alignment = Alignment(horizontal='center')

    # Форматування діаграми
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row,
                            min_col=1, max_col=(totalIntervals + 2)):
        for cell in row:
            cell.border = thin_border

    # Встановлення ширини стовпців
    columnWidth = 2
    for col in range(2, (totalIntervals + 3)):
        col_letter = get_column_letter(col)
        ws.column_dimensions[col_letter].width = columnWidth

    # Додавання товстих вертикальних ліній кожні 6 колонок (1 година)
    for blockStartCol in range(2, (totalIntervals + 3), 6):
        col_letter = get_column_letter(blockStartCol)
        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=blockStartCol)
            border_sides = cell.border
            cell.border = Border(
                left=Side(style='thick'),
                right=border_sides.right,
                top=border_sides.top,
                bottom=border_sides.bottom
            )

    # Збереження файлу
    wb.save(output_filename)
    logger.info(f"Діаграму Ганта збережено у файл: {output_filename}")

def FindOptimalLoadingDiagram(q_part_file_path, details_file_path_list, stanok_file_path, file_numbers):
    logger.info("Початок процесу знаходження оптимального розкладу")
    # Зчитування даних
    qty_dict, volume_dict = ReadBatchData(q_part_file_path)
    PartNames = qty_dict.keys()
    ProcessingTimeArr, MachineAssign, partOpsCount, partOpFileMapping = ReadProcessingData(
        details_file_path_list, PartNames, file_numbers)
    machineAvailability, machineInstances = ReadMachineData(stanok_file_path)

    # Розрахунок розкладу
    makespan, scheduleRecArray = ComputeDynamicSchedule(
        qty_dict, volume_dict, ProcessingTimeArr, MachineAssign, partOpsCount, machineAvailability, partOpFileMapping
    )

    # Генерування діаграми Ганта
    timestamp = time.strftime('%Y%m%d_%H%M%S')
    output_filename = f'GanttChart_{timestamp}.xlsx'
    DrawGanttChartTable(scheduleRecArray, makespan, machineInstances, output_filename=output_filename)
    return output_filename  # Повертаємо ім'я згенерованого файлу
